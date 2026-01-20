import os
from datetime import datetime, timedelta
from pathlib import Path
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side
import locale

def add_horizontal_line_above(worksheet, target_row, from_col, to_col):
    """Add a thin horizontal border above the specified cells."""
    thin_border = Border(top=Side(style='thin'))
    for col in range(from_col, to_col + 1):
        cell = worksheet.cell(row=target_row, column=col)
        current_border = cell.border
        cell.border = Border(
            top=Side(style='thin'),
            bottom=current_border.bottom,
            left=current_border.left,
            right=current_border.right
        )

def main():
    # Set Finnish locale for month and day names
    try:
        locale.setlocale(locale.LC_TIME, 'fi_FI.UTF-8')
    except:
        try:
            locale.setlocale(locale.LC_TIME, 'Finnish_Finland.1252')
        except:
            print("Warning: Could not set Finnish locale. Using default locale.")
    
    while True:
        # Ask for the desired year
        print("Kirjoita mille vuodelle tämä tehdään ja paina \"Enter\".")
        
        input_str = input()
        year = 0
        requested_year = datetime(1970, 1, 1)
        
        if input_str:
            try:
                year = int(input_str)
            except ValueError:
                print("\nJos tällä kertaa antaasit numeroina vuosiluvun...\n\n")
                continue
            
            if year < datetime.now().year - 2:
                print(f"\nTaitaa olla myöhästä tehdä tilikauden kirjanpitoa vuodelle {year}. Koitahan uudestaan.\n\n")
                continue
            
            try:
                requested_year = datetime(year, 1, 1)
            except Exception:
                print("\nOof, ei tollasta vuosilukua pysty edes parsimaan...")
                print("Yritäs uudestaan.\n\n")
                continue
        
        print("Kirjoita kenelle tämä tulee ja paina \"Enter\".")
        name = input()
        
        print("\nTyöstetään...\n")
        
        # Define filename and make sure the intended path for it exists
        documents_path = Path.home() / "Documents" / "Vip-Hius"
        documents_path.mkdir(parents=True, exist_ok=True)
        
        filename = f"{name}_tilikausi_{year}.xlsx"
        path = documents_path / filename
        
        if path.exists():
            print("Tälle nimelle ja vuodelle on jo tilikausi-dokumentti olemassa! \nKäytä toista vuosilukua tai nimeä, tai poista tai siirrä olemassa oleva dokumentti toiseen kansioon ennen kuin yrität uudestaan.\n\n\n")
            continue
        
        # Create Excel workbook
        workbook = Workbook()
        # Remove default sheet
        if "Sheet" in workbook.sheetnames:
            del workbook["Sheet"]
        
        # Used for finding the monthly summary cells for yearly summary page
        monthly_total_rows = {}
        
        a = 1
        b = 2
        c = 3
        d = 4
        e = 5
        f = 6
        
        current_date = requested_year
        
        # Iterate months, adding the dates, days, and formulas to the corresponding cells
        for i in range(12):
            month = current_date.strftime("%B").capitalize()
            print(month)
            worksheet = workbook.create_sheet(title=month)
            
            # Sheet title
            worksheet.cell(1, c).value = f"Vip-Hius {name}"
            worksheet.cell(1, c).alignment = Alignment(horizontal='left')
            worksheet.cell(1, c).font = Font(bold=True, italic=True, underline='single', size=15)
            worksheet.cell(2, e).value = f"{month} {year}"
            
            worksheet.cell(3, c).value = "Alv 24%"
            worksheet.cell(3, d).value = "Alv 24%"
            
            # Sheet headers
            worksheet.cell(4, a).value = "Pvm"
            worksheet.cell(4, a).font = Font(bold=True)
            worksheet.column_dimensions[chr(64 + a)].width = 4.5
            for row in worksheet.iter_rows(min_row=5, max_row=worksheet.max_row, min_col=a, max_col=a):
                for cell in row:
                    cell.alignment = Alignment(horizontal='center')
            
            worksheet.cell(4, b).value = "Päivä"
            worksheet.cell(4, b).font = Font(bold=True)
            worksheet.column_dimensions[chr(64 + b)].width = 5
            for row in worksheet.iter_rows(min_row=5, max_row=worksheet.max_row, min_col=b, max_col=b):
                for cell in row:
                    cell.alignment = Alignment(horizontal='center')
            
            worksheet.cell(4, c).value = "Työt"
            worksheet.cell(4, c).font = Font(bold=True)
            worksheet.column_dimensions[chr(64 + c)].width = 12
            
            worksheet.cell(4, d).value = "Aineet"
            worksheet.cell(4, d).font = Font(bold=True)
            worksheet.column_dimensions[chr(64 + d)].width = 12
            
            worksheet.cell(4, e).value = "Viikkokoonti"
            worksheet.cell(4, e).font = Font(bold=True)
            
            # Weekly summary headers
            worksheet.cell(5, e).value = "Työt"
            worksheet.cell(5, e).font = Font(bold=True)
            worksheet.column_dimensions[chr(64 + e)].width = 12
            
            worksheet.cell(5, f).value = "Aineet"
            worksheet.cell(5, f).font = Font(bold=True)
            worksheet.column_dimensions[chr(64 + f)].width = 12
            
            # Get number of days in the current month
            if current_date.month == 12:
                next_month = datetime(current_date.year + 1, 1, 1)
            else:
                next_month = datetime(current_date.year, current_date.month + 1, 1)
            days_in_month = (next_month - current_date).days
            
            summary_offset = 0
            row_offset = 6
            day_number_offset = 1
            row_counter = row_offset
            last_summary_row = 0
            
            summary_rows_tyot = []
            summary_rows_aineet = []
            
            day_counter = 0
            while day_counter < days_in_month:
                current_day = current_date + timedelta(days=day_counter)
                
                worksheet.cell(row_counter, a).value = day_counter + 1
                worksheet.cell(row_counter, a).font = Font(bold=True)
                worksheet.cell(row_counter, a).alignment = Alignment(horizontal='center')
                
                day_name = current_day.strftime("%A")[:2]
                worksheet.cell(row_counter, b).value = day_name
                worksheet.cell(row_counter, b).font = Font(bold=True)
                worksheet.cell(row_counter, b).alignment = Alignment(horizontal='center')
                
                worksheet.cell(row_counter, c).number_format = '0.00€'
                worksheet.cell(row_counter, c).value = 0
                worksheet.cell(row_counter, d).number_format = '0.00€'
                worksheet.cell(row_counter, d).value = 0
                
                day_counter += 1
                row_counter += 1
                
                if current_day.weekday() == 6:  # Sunday (Monday is 0, Sunday is 6)
                    # Add summary formulas for this week
                    week_length = row_counter - last_summary_row - 1
                    
                    # Työt
                    cell = worksheet.cell(row_counter, e)
                    cell.value = f"=SUM(C{row_counter - 1}:C{row_counter - week_length})"
                    cell.font = Font(bold=True)
                    cell.border = Border(
                        left=Side(style='medium'),
                        right=Side(style='medium'),
                        top=Side(style='medium'),
                        bottom=Side(style='medium')
                    )
                    cell.number_format = '0.00€'
                    summary_rows_tyot.append(f"C{row_counter}")
                    
                    # Aineet
                    cell = worksheet.cell(row_counter, f)
                    cell.value = f"=SUM(D{row_counter - 1}:D{row_counter - week_length})"
                    cell.font = Font(bold=True)
                    cell.border = Border(
                        left=Side(style='medium'),
                        right=Side(style='medium'),
                        top=Side(style='medium'),
                        bottom=Side(style='medium')
                    )
                    cell.number_format = '0.00€'
                    summary_rows_aineet.append(f"D{row_counter}")
                    
                    add_horizontal_line_above(worksheet, row_counter, 1, 4)
                    
                    last_summary_row = row_counter
                    row_counter += 1
            
            if row_counter > last_summary_row + 1:
                # Add last weekly total summary formulas
                week_length = row_counter - last_summary_row - 1
                
                # Työt
                cell = worksheet.cell(row_counter, e)
                cell.value = f"=SUM(C{row_counter - 1}:C{row_counter - week_length})"
                cell.font = Font(bold=True)
                cell.border = Border(
                    left=Side(style='medium'),
                    right=Side(style='medium'),
                    top=Side(style='medium'),
                    bottom=Side(style='medium')
                )
                cell.number_format = '0.00€'
                summary_rows_tyot.append(f"C{row_counter}")
                
                # Aineet
                cell = worksheet.cell(row_counter, f)
                cell.value = f"=SUM(D{row_counter - 1}:D{row_counter - week_length})"
                cell.font = Font(bold=True)
                cell.border = Border(
                    left=Side(style='medium'),
                    right=Side(style='medium'),
                    top=Side(style='medium'),
                    bottom=Side(style='medium')
                )
                cell.number_format = '0.00€'
                summary_rows_aineet.append(f"D{row_counter}")
                
                add_horizontal_line_above(worksheet, row_counter, 1, 4)
            
            row_counter += 2
            
            worksheet.cell(row_counter, e).value = "Kuukausikoonti"
            worksheet.cell(row_counter, e).font = Font(bold=True)
            
            row_counter += 1
            
            # Monthly summary headers
            worksheet.cell(row_counter, e).value = "Työt"
            worksheet.cell(row_counter, e).font = Font(bold=True)
            worksheet.cell(row_counter, f).value = "Aineet"
            worksheet.cell(row_counter, f).font = Font(bold=True)
            
            row_counter += 1
            
            # Add month total summary formulas
            # Työt
            tyot_refs = [f"E{ref[1:]}" for ref in summary_rows_tyot]
            cell = worksheet.cell(row_counter, e)
            cell.value = f"={'+'.join(tyot_refs)}"
            cell.font = Font(bold=True)
            cell.border = Border(
                left=Side(style='medium'),
                right=Side(style='medium'),
                top=Side(style='medium'),
                bottom=Side(style='medium')
            )
            cell.number_format = '0.00€'
            
            # Aineet
            aineet_refs = [f"F{ref[1:]}" for ref in summary_rows_tyot]
            cell = worksheet.cell(row_counter, f)
            cell.value = f"={'+'.join(aineet_refs)}"
            cell.font = Font(bold=True)
            cell.border = Border(
                left=Side(style='medium'),
                right=Side(style='medium'),
                top=Side(style='medium'),
                bottom=Side(style='medium')
            )
            cell.number_format = '0.00€'
            
            monthly_total_rows[month] = row_counter
            
            add_horizontal_line_above(worksheet, row_counter, 1, 4)
            
            row_counter += 1
            
            # Alv summary
            worksheet.cell(row_counter, e).value = "Alv 24%"
            worksheet.cell(row_counter, f).value = "Alv 24%"
            
            row_counter += 1
            worksheet.cell(row_counter, e).number_format = '0.00€'
            worksheet.cell(row_counter, e).value = f"=E{monthly_total_rows[month]}*0.24"
            worksheet.cell(row_counter, f).number_format = '0.00€'
            worksheet.cell(row_counter, f).value = f"=F{monthly_total_rows[month]}*0.24"
            
            # Move to next month
            if current_date.month == 12:
                current_date = datetime(current_date.year + 1, 1, 1)
            else:
                current_date = datetime(current_date.year, current_date.month + 1, 1)
        
        # Add year summary sheet
        year_summary_worksheet = workbook.create_sheet(title="Vuosikoonti")
        
        # Year summary sheet title
        year_summary_worksheet.cell(1, b).value = f"Vip-Hius {name}"
        year_summary_worksheet.cell(1, b).alignment = Alignment(horizontal='left')
        year_summary_worksheet.cell(1, b).font = Font(bold=True, italic=True, underline='single', size=15)
        
        year_summary_worksheet.cell(4, a).value = f"Vuosikoonti {year}"
        year_summary_worksheet.cell(4, a).font = Font(bold=True, size=12)
        year_summary_worksheet.column_dimensions[chr(64 + a)].width = 15
        
        year_summary_worksheet.cell(6, b).value = "Työt"
        year_summary_worksheet.cell(6, b).font = Font(bold=True)
        year_summary_worksheet.column_dimensions[chr(64 + b)].width = 12
        
        year_summary_worksheet.cell(6, c).value = "Aineet"
        year_summary_worksheet.cell(6, c).font = Font(bold=True)
        year_summary_worksheet.column_dimensions[chr(64 + c)].width = 12
        
        row = 7
        
        # Year summary
        for month_name, month_row in monthly_total_rows.items():
            add_horizontal_line_above(year_summary_worksheet, row, 1, 3)
            
            # month name
            year_summary_worksheet.cell(row, a).value = month_name
            year_summary_worksheet.cell(row, a).font = Font(bold=True)
            
            # Työt total
            year_summary_worksheet.cell(row, b).value = f"={month_name}!E{month_row}"
            year_summary_worksheet.cell(row, b).font = Font(bold=True)
            year_summary_worksheet.cell(row, b).number_format = '0.00€'
            
            # Aineet total
            year_summary_worksheet.cell(row, c).value = f"={month_name}!F{month_row}"
            year_summary_worksheet.cell(row, c).font = Font(bold=True)
            year_summary_worksheet.cell(row, c).number_format = '0.00€'
            
            row += 1
        
        add_horizontal_line_above(year_summary_worksheet, row, a, c)
        
        year_totals_row = 21
        
        # Year totals
        year_summary_worksheet.cell(20, a).value = "Koko Vuosi"
        year_summary_worksheet.cell(20, a).font = Font(bold=True)
        year_summary_worksheet.cell(20, b).value = "Työt"
        year_summary_worksheet.cell(20, b).font = Font(bold=True)
        year_summary_worksheet.cell(20, c).value = "Aineet"
        year_summary_worksheet.cell(20, c).font = Font(bold=True)
        
        year_summary_worksheet.cell(year_totals_row, b).value = "=SUM(B7:B18)"
        year_summary_worksheet.cell(year_totals_row, b).font = Font(bold=True)
        year_summary_worksheet.cell(year_totals_row, b).number_format = '0.00€'
        
        year_summary_worksheet.cell(year_totals_row, c).value = "=SUM(C7:C18)"
        year_summary_worksheet.cell(year_totals_row, c).font = Font(bold=True)
        year_summary_worksheet.cell(year_totals_row, c).number_format = '0.00€'
        
        # Alv 24%
        year_summary_worksheet.cell(23, b).value = "Alv 24%"
        year_summary_worksheet.cell(23, c).value = "Alv 24%"
        
        year_summary_worksheet.cell(24, b).number_format = '0.00€'
        year_summary_worksheet.cell(24, b).value = f"=B{year_totals_row}*0.24"
        year_summary_worksheet.cell(24, c).number_format = '0.00€'
        year_summary_worksheet.cell(24, c).value = f"=C{year_totals_row}*0.24"
        
        # Save the workbook
        workbook.save(path)
        
        print(f"Tiedosto tallennettu sijaintiin\"{path}\"")
        print("Paina \"Q\" sulkeaksesi sovelluksen, tai \"Enter\" tehdäksesi uuden dokumentin.")
        
        # Ask user for quit or another round
        user_input = input()
        
        if user_input.upper() == "Q":
            return
        else:
            continue

if __name__ == "__main__":
    main()
